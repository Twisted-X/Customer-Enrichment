"""
Configuration for Twisted X Scraper API
"""
import os
import csv
import logging

log = logging.getLogger(__name__)

# Playwright settings
HEADLESS = True
TIMEOUT_MS = 30000  # 30 seconds page load timeout

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")

# Fallback retailer URLs (used when CSV is not available)
RETAILER_URLS = [
    "https://blueribboncountrystore.com/",
    "https://bluestemfarmandranch.com",
    "https://bobsmerch.com",
    "https://bomgaars.com",
    "https://bonypony.com",
    "https://bootamerica.com/",
    "https://bootbarn.com/",
    "https://bootbox.com/",
    "https://bootconnection.com",
    "https://bootcountryonline.com/",
    "https://bootjack.com/",
    "https://bootleggersfootwear.com/",
    "https://bootsandjeans.net",
    "https://bootsandmore.net/",
    "https://bootsblingcowboythings.com",
    "https://bootsforlessomaha.com",
    "https://bootsnbritchestx.com/",
    "https://bootvil.square.site/",
    "https://bootybrotherswestern.com",
    "https://bostonshoestoboots.com",
    "https://botasrojerowesternwear.com",
    "https://brayssaddlery.com",
    "https://brianfarmservice.com",
    "https://broadrivermercantile.com",
    "https://bromarion.com/",
    "https://broncotrading.co",
    "https://brownsshoefitco.com",
    "https://brownsshoes.com",
    "https://brutesafetywear.com",
    "https://buchheits.com/",
    "https://bullriderharlingen.com/",
    "https://burgershoes.com",
    "https://burkhartzmeyershoes.net",
    "https://burlapbagcos.com",
    "https://burrisfarmhome.com",
    "https://byrdswesternstore.com",
    "https://calfarley.org",
    "https://calliekays.com",
    "https://calranch.com",
    "https://cam-safety.com/",
    "https://capehornwesternwear.com/",
    "https://capitalcityshoes.com",
    "https://carrollsbootcountry.com",
    "https://casaraulww.com/",
    "https://cascademerchantile.com",
    "https://catalenahatters.com",
    "https://cattlemanswesternwear.com",
]


def get_retailer_name(url: str) -> str:
    """Extract retailer name from URL (e.g. 'https://www.bootbarn.com/' -> 'bootbarn')"""
    from urllib.parse import urlparse
    parsed = urlparse(url)
    domain = parsed.netloc.replace("www.", "")
    name = domain.split(".")[0]
    return name


# ─── Twisted X SKU Fingerprint Database ─────────────────────────────────────
# Loaded once at import time. Used by /api/check for verification.
# Covers: Twisted X, Black Star, Tamarindo Footwear, Wrangler Footwear

# Source files in DATA_DIR (overridable via env vars for non-default deployments)
SKU_XLSX_FILENAME = os.getenv("SKU_XLSX_FILENAME", "twisted_x_skus_v107.xlsx")
SKU_CSV_FILENAME = os.getenv("SKU_CSV_FILENAME", "twisted_x_sku.csv")

# Minimum count below which the SKU database is considered broken; raises at
# import time so a misconfigured deployment fails fast instead of silently
# returning empty matches.
MIN_EXPECTED_STYLE_CODES = int(os.getenv("MIN_EXPECTED_STYLE_CODES", "1000"))


def _load_style_codes() -> set:
    """
    Load unique ParentStyle codes from the full SKU xlsx (all sheets)
    with CSV fallback. These are product fingerprints for verification.
    Returns a set of uppercase style codes for O(1) lookup.
    """
    styles: set = set()

    xlsx_path = os.path.join(DATA_DIR, SKU_XLSX_FILENAME)
    if os.path.exists(xlsx_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            sheet_count = len(wb.sheetnames)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ps_idx = -1
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0:
                        headers = list(row)
                        ps_idx = headers.index("ParentStyle") if "ParentStyle" in headers else -1
                        continue
                    if ps_idx >= 0 and row[ps_idx]:
                        style = str(row[ps_idx]).strip()
                        if ":" not in style and len(style) >= 4:
                            styles.add(style.upper())
            wb.close()
            log.info("Loaded %d style codes from xlsx (%d sheets)", len(styles), sheet_count)
            return styles
        except Exception as e:
            log.warning("xlsx load failed (%s), falling back to CSV", e)

    sku_csv = os.path.join(DATA_DIR, SKU_CSV_FILENAME)
    if os.path.exists(sku_csv):
        try:
            with open(sku_csv, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    style = row.get("ParentStyle", "").strip()
                    if style and ":" not in style and len(style) >= 4:
                        styles.add(style.upper())
            log.info("Loaded %d style codes from CSV (fallback)", len(styles))
        except Exception as e:
            log.error("Failed to load SKU file %s: %s", sku_csv, e)
    else:
        log.error("No SKU files found in %s", DATA_DIR)

    return styles


TX_STYLE_CODES = _load_style_codes()

if len(TX_STYLE_CODES) < MIN_EXPECTED_STYLE_CODES:
    raise RuntimeError(
        f"SKU database failed to load: got {len(TX_STYLE_CODES)} style codes, "
        f"expected at least {MIN_EXPECTED_STYLE_CODES}. "
        f"Check that {SKU_XLSX_FILENAME} or {SKU_CSV_FILENAME} exists in {DATA_DIR}."
    )
